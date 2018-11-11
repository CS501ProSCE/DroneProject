"""
CS 50100 Group 16
Written by Tom shaw

Python script to create data files from DataDefinition.xlsx data catalog
1. Modify DataDefinitions.xlsx to define your flight test data
2. Modify MavlinkParameters.py to define the data parameters you want to extract
3. Ensure all your data file references (*.csv) files are in your run directory
4. execute

Output is:
    *.txt file for each mavlink parameters for train and test data series
    train_labels.txt file that defines labels for all training data
    test_labels.txt file that defines labels for all test data

"""


# System files
import openpyxl as op
import numpy as np
import datetime
import os

# Custom files
from MavlinkParameters import mavlink_types
from MavlinkParameters import mavlink_param
from MavlinkParameters import mavlink_index
from MavlinkParameters import mavlink_rate

dataset = 'Data2/'

#Import Data Definition catalog
wbpath = dataset + 'DataDefinitions.xlsx'
book = op.load_workbook(wbpath) #This is your data definition spreadsheet
sheet = book.active
maxRow = sheet.max_row #get number of test slices to import

#initialize 
filelist = []
labellist = []
testtypelist= []
starttimelist = []
timelist = []

#read in Data Definition tests
for row in range(2,maxRow+1):
    filelist.append(sheet.cell(row=row,column=2).value)
    labellist.append(sheet.cell(row=row,column=3).value)
    testtypelist.append(sheet.cell(row=row,column=4).value)
    starttimelist.append(sheet.cell(row=row,column=5).value)
    timelist.append(sheet.cell(row=row,column=6).value)
    
print('DataDefinition has %i data samples to process' %(len(filelist)))
       
#Loop through each test case, read data, write out time slice
for testnum in range(len(filelist)):

    #re-initialize list stucture
    mavlink_data = []
    for typ in range(len(mavlink_types)):
        typdata = []
        for _ in range(len(mavlink_param[typ])):
            typdata.append([]) 
        mavlink_data.append(typdata)    
    
    #Decompose test 
    startt = datetime.datetime.strptime(starttimelist[testnum].replace("T"," "), '%Y-%m-%d %H:%M:%S.%f')
    testime = int(timelist[testnum])
            
    #read raw data file
    filelines = []
    datapath = dataset + filelist[testnum]
    with open(datapath) as f: 
        for line in f:
            filelines.append(line.split(','))
    f.close()
    
    #read all data from file into parameter lists
    for row in filelines:
        for typ in range(len(mavlink_types)):
            if(str(row[9]) == mavlink_types[typ]):          
                for param in range(len(mavlink_param[typ])):
                    if(param==0):   #leave timestamp as string
                        mavlink_data[typ][param].append(row[mavlink_index[typ][param]-1])
                    else:           #convert all other to float
                        mavlink_data[typ][param].append(float(row[mavlink_index[typ][param]-1]))          

    #Loop through test case to find start and stop iteration
    for typ in range(len(mavlink_types)):
        testit = int(testime * mavlink_rate[typ]) #find number of iterations
        for ndata in range(len(mavlink_data[typ][0])):
            timestamp = mavlink_data[typ][0][ndata]
            timestamp = timestamp.replace("T"," ") #must remove extra 'T' in Mavlink timestamp
            time = datetime.datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S.%f')                
            if(time>startt):
                data_it_start = ndata
                if(len(mavlink_data[typ][0]) < (data_it_start + testit)):
                    print(data_it_start)
                    print(datetime.timedelta(seconds=testime))
                    raise ValueError("ERROR: Test number " + str(testnum+1) + "has insufficient length for: " + str(mavlink_types[typ]))              
                break
            
        for param in range(len(mavlink_param[typ])):
            #write to data file (X axis)
            filename = dataset + testtypelist[testnum].lower() + "_" + mavlink_types[typ] + "_" + mavlink_param[typ][param] + ".txt"
            if os.path.exists(filename):
                append_write  = 'a'
            else:
                append_write= 'w'
            file = open(filename, append_write)
            file.write(' '.join(map(str,mavlink_data[typ][param][data_it_start:data_it_start + testit])))
            file.write("\n")
            file.close()  
            
    #write to label file (Y axis) for each test
    labelfilename = dataset + testtypelist[testnum].lower() + "_labels.txt"
    if os.path.exists(labelfilename):
        append_write  = 'a'
    else:
        append_write= 'w'
    file = open(labelfilename, append_write)
    file.write(str(labellist[testnum]))
    file.write("\n")
    file.close()              
          
print('Data generation is complete')

    
    
