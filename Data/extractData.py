
"""
CS 50100 Group 16
Written by Thomas Shaw

Python script to extract raw drone data from Mission Planner output

1. input csv file name
2. determine if you want to cut sample 
3. xlsx file will be output as "Data_filenmame.xlsx"

"""

import openpyxl as op
import numpy as np
import datetime
import os

from MavlinkParameters import mavlink_types
from MavlinkParameters import mavlink_param
from MavlinkParameters import mavlink_index

#change this to name of file - MUST BE in XLS format to work with openpyxl!!
name = "2018-10-07 13-49-59.csv"

#Define file options
createspreadsheet = False #define if you want MAVLINK data in spreadsheet form

cuttime = True  #define True if you want to cut to start and stop times
starttime = "2018-10-07T13:51:42.813"  #Data will start at this time
endtime = "2018-10-07T13:51:47.972" #Data will end at this time

appenddata = True #This will append a data series, as a row, to requested file
appendname = "test" #define as "train" or "test"


if(cuttime):
    starttime = starttime.replace("T"," ")
    endtime = endtime.replace("T"," ")
    startt = datetime.datetime.strptime(starttime, '%Y-%m-%d %H:%M:%S.%f')
    endt = datetime.datetime.strptime(endtime, '%Y-%m-%d %H:%M:%S.%f')
    print("This will process file %s"%(name))
    print("The data will start at %s"%(startt))
    print("The data will end at %s"%(endt))
    print("\n")



#error check
for typ in range(len(mavlink_types)):
    if(len(mavlink_param[typ]) != len(mavlink_index[typ])):
        raise ValueError("ERROR: mavlink type " + mavlink_types[typ] + " has mismatched index and param size")

#Read data from file
        
#Create list stucture
mavlink_data = []
for typ in range(len(mavlink_types)):
    typdata = []
    for _ in range(len(mavlink_param[typ])):
        typdata.append([]) 
    mavlink_data.append(typdata)

#read raw data file
filelist = []
with open(name) as f: 
    for line in f:
        filelist.append(line.split(','))
f.close()

#read data into parameter lists
for row in filelist:
    for typ in range(len(mavlink_types)):
        if(str(row[9]) == mavlink_types[typ]):
            if(cuttime):
                timestamp = row[0]
                timestamp = timestamp.replace("T"," ")
                rowtime = datetime.datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S.%f')
                if((rowtime>startt)and(rowtime<endt)):
                    for param in range(len(mavlink_param[typ])):
                        if(param==0):   #leave timestapm as string
                            mavlink_data[typ][param].append(row[mavlink_index[typ][param]-1])
                        else:           #convert all other to float
                            mavlink_data[typ][param].append(float(row[mavlink_index[typ][param]-1]))
                    break                      
            else:
                for param in range(len(mavlink_param[typ])):
                    if(param==0):   #leave timestapm as string
                        mavlink_data[typ][param].append(row[mavlink_index[typ][param]-1])
                    else:           #convert all other to float
                        mavlink_data[typ][param].append(float(row[mavlink_index[typ][param]-1]))
                break                    

if(createspreadsheet):
    #Create new worksheet and output data
    newbook = op.Workbook()
    newbook.remove_sheet(newbook.active)
    
    ws = []
    for typ in range(len(mavlink_types)):
        ws.append(newbook.create_sheet(mavlink_types[typ]))    
        for param in range(len(mavlink_param[typ])):
            ws[typ].cell(row=1,column=param+1).value = mavlink_param[typ][param]
            for row in range(len(mavlink_data[typ][param])):
                ws[typ].cell(row=row+2,column=param+1).value = mavlink_data[typ][param][row]      
    
    newname =  "Data_" + name.replace("csv","xlsx")    
    newbook.save(newname)


if(appenddata):
#open data file and write to
    for typ in range(len(mavlink_types)):
            for param in range(len(mavlink_param[typ])):
                filename = appendname + "_" + mavlink_types[typ] + "_" + mavlink_param[typ][param] + ".txt"
                if os.path.exists(filename):
                    append_write  = 'a'
                else:
                    append_write= 'w'
                file = open(filename, append_write)
                file.write(' '.join(map(str,mavlink_data[typ][param])))
                file.write("\n")
                file.close()
                    
    
#calculate statistics
for typ in range(len(mavlink_types)):
    firsttime = mavlink_data[typ][0][0]
    firsttime = firsttime.replace("T"," ")
    lasttime = mavlink_data[typ][0][-1]
    lasttime = lasttime.replace("T"," ")
    firstT = datetime.datetime.strptime(firsttime, '%Y-%m-%d %H:%M:%S.%f')
    lastT = datetime.datetime.strptime(lasttime, '%Y-%m-%d %H:%M:%S.%f')
    timediff = lastT - firstT
    rate = len(mavlink_data[typ][0]) / (timediff.total_seconds())
    print("Parameter %s has a rate of %f" % (mavlink_types[typ],rate))
