
"""
CS 50100 Group 16
Written by Thomas Shaw

Python script to extract raw drone data from Mission Planner output

1. input csv file name
2. determine if you want to cut sample 
3. xlsx file will be output as "Data_filenmame.xlsx"

"""

import openpyxl as op
import numpy as np
import datetime
import os

from MavlinkParameters import mavlink_types
from MavlinkParameters import mavlink_param
from MavlinkParameters import mavlink_index

#change this to name of file - MUST BE in XLS format to work with openpyxl!!
path = 'Data2/'
name = "2018-11-08 16-25-11.csv"
fullpath = path + name

#Define file options
createspreadsheet = True #define if you want MAVLINK data in spreadsheet form

cuttime = True  #define True if you want to cut to start and stop times
starttime = "2018-11-08T16:26:31.350"  #Data will start at this time
endtime = "2018-11-08T16:31:27.511" #Data will end at this time


if(cuttime):
    starttime = starttime.replace("T"," ")
    endtime = endtime.replace("T"," ")
    startt = datetime.datetime.strptime(starttime, '%Y-%m-%d %H:%M:%S.%f')
    endt = datetime.datetime.strptime(endtime, '%Y-%m-%d %H:%M:%S.%f')
    print("This will process file %s"%(fullpath))
    print("The data will start at %s"%(startt))
    print("The data will end at %s"%(endt))
    print("\n")



#error check
for typ in range(len(mavlink_types)):
    if(len(mavlink_param[typ]) != len(mavlink_index[typ])):
        raise ValueError("ERROR: mavlink type " + mavlink_types[typ] + " has mismatched index and param size")

#Read data from file
        
#Create list stucture
mavlink_data = []
for typ in range(len(mavlink_types)):
    typdata = []
    for _ in range(len(mavlink_param[typ])):
        typdata.append([]) 
    mavlink_data.append(typdata)

#read raw data file
filelist = []
with open(fullpath) as f: 
    for line in f:
        filelist.append(line.split(','))
f.close()

#read data into parameter lists
for row in filelist:
    for typ in range(len(mavlink_types)):
        if(str(row[9]) == mavlink_types[typ]):
            if(cuttime):
                timestamp = row[0]
                timestamp = timestamp.replace("T"," ")
                rowtime = datetime.datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S.%f')
                if((rowtime>startt)and(rowtime<endt)):
                    for param in range(len(mavlink_param[typ])):
                        if(param==0):   #leave timestapm as string
                            mavlink_data[typ][param].append(row[mavlink_index[typ][param]-1])
                        else:           #convert all other to float
                            mavlink_data[typ][param].append(float(row[mavlink_index[typ][param]-1]))
                    break                      
            else:
                for param in range(len(mavlink_param[typ])):
                    if(param==0):   #leave timestapm as string
                        mavlink_data[typ][param].append(row[mavlink_index[typ][param]-1])
                    else:           #convert all other to float
                        mavlink_data[typ][param].append(float(row[mavlink_index[typ][param]-1]))
                break                    

if(createspreadsheet):
    #Create new worksheet and output data
    newbook = op.Workbook()
    newbook.remove_sheet(newbook.active)
    
    ws = []
    for typ in range(len(mavlink_types)):
        ws.append(newbook.create_sheet(mavlink_types[typ]))    
        for param in range(len(mavlink_param[typ])):
            ws[typ].cell(row=1,column=param+1).value = mavlink_param[typ][param]
            for row in range(len(mavlink_data[typ][param])):
                ws[typ].cell(row=row+2,column=param+1).value = mavlink_data[typ][param][row]      
    
    newname =  path + "Data_" + name.replace("csv","xlsx")    
    newbook.save(newname)

    
#calculate statistics
for typ in range(len(mavlink_types)):
    firsttime = mavlink_data[typ][0][0]
    firsttime = firsttime.replace("T"," ")
    lasttime = mavlink_data[typ][0][-1]
    lasttime = lasttime.replace("T"," ")
    firstT = datetime.datetime.strptime(firsttime, '%Y-%m-%d %H:%M:%S.%f')
    lastT = datetime.datetime.strptime(lasttime, '%Y-%m-%d %H:%M:%S.%f')
    timediff = lastT - firstT
    rate = len(mavlink_data[typ][0]) / (timediff.total_seconds())
    print("Parameter %s has a rate of %f" % (mavlink_types[typ],rate))
